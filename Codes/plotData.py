import os,re,numpy as np
import matplotlib.pyplot as plt
import pandas as pd  # 引入 pandas 库以便处理和导出数据
import scipy.io

from openpyxl import load_workbook

import matplotlib
matplotlib.use('TkAgg')  # 指定使用 Tk 后端，或根据需要选择合适的后端

# 设置中文字体为 Microsoft YaHei
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False  # 防止负号显示问题

# Define the number of experiments
num_experiments = 4

base_folder = 'D:/Documents/Peng/EEG/Datasets/CMCresult_lowerLimb/崔'


# Two types of motion: 'a' and 'b'
motion = 'a'
# Store frequency domain CMC data
EEGsch_name = ['C3','C4']
EMGsch_name = ['MG','RF']
#不是双向
CMC_num = len(EEGsch_name)*len(EMGsch_name)
subject_means = np.zeros((CMC_num,num_experiments, 5))

# Add vertical lines and text labels for frequency band divisions
alpha_freq = np.array([8,13])  # α frequency band
beta_freq = np.array([13,30])  # β frequency band
gamma_freq = np.array([30,49])  # γ frequency band
delta_freq = np.array([1,4])  # δ frequency band
theta_freq = np.array([4,8])  # θ frequency band

CMC = {}
Freq = np.arange(1,50,0.196)
for name1 in EEGsch_name:
    for name2 in EMGsch_name:
        CMC['wcohere_'+name2+'_'+name1] = {} 

# 使用布尔索引选择满足条件的索引——————左闭右开
alpha_indices = np.where((Freq >= np.min(alpha_freq)) & (Freq < np.max(alpha_freq)))[0]
beta_indices = np.where((Freq >= np.min(beta_freq)) & (Freq < np.max(beta_freq)))[0]
gamma_indices = np.where((Freq >= np.min(gamma_freq)) & (Freq < np.max(gamma_freq)))[0]
delta_indices = np.where((Freq >= np.min(delta_freq)) & (Freq < np.max(delta_freq)))[0]
theta_indices = np.where((Freq >= np.min(theta_freq)) & (Freq < np.max(theta_freq)))[0]

print(alpha_indices)
files_in_folder = []
for file_name in os.listdir(base_folder):
    file_path = os.path.join(base_folder, file_name)

    # 筛选出 .mat 文件
    if file_name.endswith(".mat") and os.path.isfile(file_path):
        # 提取前缀和后缀数字
        prefix_match = re.match(r"^(\d+)_", file_name)  # 匹配前缀数字
        # suffix_match = re.search(r"_(\\d+)\\.", file_name)  # 匹配后缀数字
        # print(prefix_match)
        # if prefix_match or suffix_match:
        if prefix_match:
            prefix_number = int(prefix_match.group(1))
            # suffix_number = int(suffix_match.group(1))
            files_in_folder.append({
                "file_path": file_path,
                "prefix": prefix_number,
                # "suffix": suffix_number
            })
print(files_in_folder)
# 按后缀数字排序文件
files_in_folder.sort(key=lambda x: x["prefix"])

for data_file in files_in_folder:
    # Read the experiment data
    data = scipy.io.loadmat(data_file['file_path'])
    test = data_file['prefix']
    data = data['results']
    # Calculate frequency domain CMC (sum along rows)
    for name1 in EEGsch_name:
        for name2 in EMGsch_name:
            #1代表对第二个维度处理
            CMC['wcohere_'+name2+'_'+name1]['s0'+str(data_file['prefix'])] = np.mean(data['wcohere_'+name1+'_'+name2], axis=1)
    
    for k in range(CMC_num):
        alpha_mean = np.mean(CMC[list(CMC.keys())[k]]['s0'+str(test)][0][alpha_indices])
        beta_mean = np.mean(CMC[list(CMC.keys())[k]]['s0'+str(test)][0][beta_indices])
        gamma_mean = np.mean(CMC[list(CMC.keys())[k]]['s0'+str(test)][0][gamma_indices])
        delta_mean = np.mean(CMC[list(CMC.keys())[k]]['s0'+str(test)][0][delta_indices])
        theta_mean = np.mean(CMC[list(CMC.keys())[k]]['s0'+str(test)][0][theta_indices])
        subject_means[k, test-1, :] = [alpha_mean, beta_mean, gamma_mean, delta_mean, theta_mean]
        
print(subject_means[0])

# plot_path = f'D:/Documents/Peng/zhengda/CMCresult_lowerLimb/CMC_Freq/subj{subject_id}/{motion}/'
plot_path = f'{base_folder}/plots/'
if not os.path.exists(plot_path):
    os.makedirs(plot_path)
        # Plot the bar chart
for j in range(CMC_num):
    plt.figure()
    x = np.arange(5)
    heights = subject_means [j,:,:]

    for i, height in enumerate(heights):
        x = np.arange(5) + i*0.2  # 调整x坐标，使每个主题的柱形图错开
        plt.bar(x, height, width=0.2, label=f's0{i+1}')  # 绘制每个主题的柱形图，并添加标签

    base_name = os.path.basename(base_folder)
    tit2 = f'Freq_{base_name}-{motion}-{list(CMC.keys())[j]}'
    plt.title(tit2)
    plt.xlabel('Freq')
    plt.ylabel('CMC')
    plt.legend()
    plt.xticks(np.arange(5), ['α', 'β', 'γ', 'δ', 'θ'])
    plt.savefig(plot_path + f'{tit2}.png')

    plt.show()
    # plt.figure()
    # temp_max = np.zeros((4, 1))
    # for i in range(1, num_experiments + 1):
    #     x = Freq
    #     y = CMC[list(CMC.keys())[j]]['s0'+str(i)][0]
    #     plt.plot(x, y, label=f's{i:d}')
    #     temp_max[i-1,0] = np.max(y)
    # Max = np.max(temp_max)
    # plt.axvline(x=alpha_freq[0], linestyle='--', color='red')
    # plt.axvline(x=alpha_freq[-1], linestyle='--', color='red')
    # plt.text(alpha_freq[0], Max, 'α', ha='left', va='bottom')

    # plt.axvline(x=beta_freq[0], linestyle='--', color='blue')
    # plt.axvline(x=beta_freq[-1], linestyle='--', color='blue')
    # plt.text(beta_freq[0], Max, 'β', ha='left', va='bottom')

    # plt.axvline(x=gamma_freq[0], linestyle='--', color='green')
    # plt.axvline(x=gamma_freq[-1], linestyle='--', color='green')
    # plt.text(gamma_freq[0], Max, 'γ', ha='left', va='bottom')

    # plt.axvline(x=delta_freq[0], linestyle='--', color='orange')
    # plt.axvline(x=delta_freq[-1], linestyle='--', color='orange')
    # plt.text(delta_freq[0], Max, 'δ', ha='left', va='bottom')

    # plt.axvline(x=theta_freq[0], linestyle='--', color='purple')
    # plt.axvline(x=theta_freq[-1], linestyle='--', color='purple')
    # plt.text(theta_freq[0], Max, 'θ', ha='left', va='bottom')
    
    # tit1 = f'Freq-subj{base_name}-{motion}-{list(CMC.keys())[j]}'
    # plt.title(f'Freq-subj{base_name}-{motion}-{list(CMC.keys())[j]}')
    # plt.xlabel('Frequency')
    # plt.ylabel('CMC')
    # plt.xlim(0, 50)
    # plt.legend()
    # plt.show()
    # plt.savefig(plot_path + f'{tit1}.png')
            

# save to excel


titles = ['名字']
for i, name1 in enumerate(EEGsch_name):
    for j, name2 in enumerate(EMGsch_name):
        title_name = f'{name2}-{name1}'
        
        # 创建标题行
        titles.append(title_name)
        
parent_folder = os.path.dirname(base_folder)
for freq_band, freq_id in zip(['α', 'β', 'γ', 'δ', 'θ'], range(5)):
    print(freq_id)
    excel_path = f"{parent_folder}/{motion}_{freq_band}.xlsx"
    # 判断文件是否存在，如果不存在则创建
    if not os.path.exists(excel_path):
        # 创建一个Excel文件
        with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
            # 创建五个频段对应的sheet
            for test_id in range(num_experiments):
                
                # 初始化空数据框，列为标题行
                df = pd.DataFrame(columns=titles)
                
                # 将DataFrame写入对应的sheet
                # df.to_excel(writer, sheet_name=test_id, index=False)
                df.to_excel(writer, sheet_name=f'test_{test_id}', index=False)

    # 打开excel并写入数据
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
        # workbook = load_workbook(excel_path)
        # writer.book = workbook
        # writer.sheets = {ws.title: ws for ws in workbook.worksheets}
        # 按频率区间填充数据
        for test_id in range(num_experiments):
            # 获取对应的sheet
            sheet_name = f'test_{test_id}'
            
            data_row = [f'{base_name}', subject_means[0][test_id][freq_id], subject_means[1][test_id][freq_id], subject_means[2][test_id][freq_id], subject_means[3][test_id][freq_id]]
            # 将数据填充到Excel文件中的相应sheet
            df = pd.DataFrame([data_row], columns=titles)
            
            #  # 检查目标 sheet 的行数
            # if sheet_name in writer.sheets:
            #     start_row = writer.sheets[sheet_name].max_row
            # else:
            #     start_row = 0
                
            df.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=writer.sheets[sheet_name].max_row)
